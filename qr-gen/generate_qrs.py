"""Generate 999 branded QR PNGs + CSVs + XLSX for Yolkshire's Golden Yolk Loyalty Program.

Output:
  ./qrs/YSLC001.png ... ./qrs/YSLC999.png     — 999 QR images
  ./qrs.csv                                   — full CSV with qr_url column
  ./qrs-indesign.csv                          — InDesign Data Merge CSV (image column @-prefixed)
  ./qrs-chunk-{1..4}.csv                      — 300-row CSV chunks (URL-based, for InDesign/Marq)
  ./qrs-chunk-{1..4}.xlsx                     — 300-row XLSX chunks with QR images
                                                 EMBEDDED in cells (for Canva Bulk Create)

Install once: pip install "qrcode[pil]" openpyxl
Run:          python generate_qrs.py
Re-running is safe — existing PNGs are skipped; CSVs/XLSX always re-emit.
"""
import csv
from pathlib import Path
from qrcode.main import QRCode
from qrcode.constants import ERROR_CORRECT_H
from qrcode.image.styledpil import StyledPilImage
from qrcode.image.styles.moduledrawers.pil import RoundedModuleDrawer
from qrcode.image.styles.colormasks import SolidFillColorMask
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

BASE_URL     = "https://yolkshire-pune.github.io/loyalty-card-tracker/?id="
QR_HOST_URL  = "https://yolkshire-golden-yolk-qr.netlify.app/"   # public host for QR PNGs (Canva Bulk Create)
HERE         = Path(__file__).parent
OUT_DIR      = HERE / "qrs"
LOGO         = HERE.parent / "yolkshire-new-logo-fhd.png"   # repo-root logo, not duplicated

# Brand palette (RGB)
GOLD       = (245, 185, 52)   # #f5b934 — QR modules (foreground)
DARK_GREEN = (52,  87, 54)    # #345736 — background

OUT_DIR.mkdir(exist_ok=True)
generated, skipped = 0, 0

# Standard CSV: works for Canva Bulk Create, Figma plugins, any spreadsheet import.
std_csv = HERE / "qrs.csv"
# InDesign Data Merge CSV: image column header must start with @ and contain a path
# relative to the CSV's location (InDesign looks up the image at merge time).
id_csv  = HERE / "qrs-indesign.csv"

with std_csv.open('w', newline='', encoding='utf-8') as fs, \
     id_csv.open('w', newline='', encoding='utf-8') as fi:
    w_std = csv.writer(fs); w_std.writerow(['card_id', 'qr_filename', 'qr_url', 'url'])
    w_id  = csv.writer(fi); w_id.writerow(['card_id', '@qr_image', 'url'])

    for i in range(1, 1000):                       # YSLC001 … YSLC999
        card_id  = f"YSLC{i:03d}"
        filename = f"{card_id}.png"
        url      = BASE_URL + card_id
        out_path = OUT_DIR / filename

        if out_path.exists():
            skipped += 1
        else:
            qr = QRCode(error_correction=ERROR_CORRECT_H, box_size=12, border=2)
            qr.add_data(url)
            qr.make(fit=True)
            kwargs = {
                "image_factory": StyledPilImage,
                "module_drawer": RoundedModuleDrawer(),
                "color_mask":    SolidFillColorMask(back_color=DARK_GREEN, front_color=GOLD),
            }
            if LOGO.exists():
                kwargs["embeded_image_path"] = str(LOGO)
            qr.make_image(**kwargs).save(out_path)
            generated += 1
            if generated % 100 == 0:
                print(f"  ...{generated} new QRs generated")

        qr_url = QR_HOST_URL + filename
        w_std.writerow([card_id, filename, qr_url, url])
        w_id.writerow([card_id, f"qrs/{filename}", url])

print(f"Done — {generated} new, {skipped} skipped. PNGs: {OUT_DIR.resolve()}")
print(f"       CSVs: {std_csv.name}, {id_csv.name}")

# Canva Bulk Create caps at 300 rows per run — emit chunked copies of qrs.csv.
CHUNK_SIZE = 300
with std_csv.open(encoding='utf-8') as f:
    reader = csv.reader(f)
    header = next(reader)
    rows = list(reader)

for i, start in enumerate(range(0, len(rows), CHUNK_SIZE), 1):
    chunk_rows = rows[start:start + CHUNK_SIZE]

    # CSV chunk (URL-based — works for InDesign/Marq; Canva treats URLs as plain text)
    chunk_csv = HERE / f"qrs-chunk-{i}.csv"
    with chunk_csv.open('w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(chunk_rows)

    # XLSX chunk with embedded QR images in cells — required by Canva Bulk Create
    wb = Workbook()
    ws = wb.active
    ws.title = "QRs"
    ws.append(['card_id', 'qr', 'qr_url', 'url'])
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16      # cell wide enough for the embedded image
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    for r, row in enumerate(chunk_rows, start=2):
        card_id, qr_filename, qr_url, url = row
        ws.cell(row=r, column=1, value=card_id)
        ws.cell(row=r, column=3, value=qr_url)
        ws.cell(row=r, column=4, value=url)
        ws.row_dimensions[r].height = 75      # ~100 px, fits the QR
        img = XLImage(str(OUT_DIR / qr_filename))
        img.width, img.height = 90, 90        # sized within the cell, not floating
        img.anchor = f'B{r}'
        ws.add_image(img)
    chunk_xlsx = HERE / f"qrs-chunk-{i}.xlsx"
    wb.save(chunk_xlsx)

    print(f"  chunk {i}: rows {start+1}-{start+len(chunk_rows)} -> {chunk_csv.name} + {chunk_xlsx.name}")
